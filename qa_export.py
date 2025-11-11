#!/usr/bin/env python3
"""QA validation of exported spreadsheet"""
import sqlite3
from pathlib import Path
from openpyxl import load_workbook

EXPORT_PATH = Path(__file__).parent / "Vehicles_Tobruk_Torch_Export.xlsx"
DB_PATH = Path(__file__).parent / "database" / "master_database.db"

# Load spreadsheet
wb = load_workbook(EXPORT_PATH)
ws = wb.active

# Load database
conn = sqlite3.connect(DB_PATH, timeout=60)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()

print("QA VALIDATION - First 5 Vehicles")
print("=" * 80)

for row_num in range(2, 7):
    excel_name = ws.cell(row_num, 1).value
    excel_move_off = ws.cell(row_num, 2).value
    excel_move_road = ws.cell(row_num, 3).value
    excel_armor_f = ws.cell(row_num, 5).value
    excel_armor_s = ws.cell(row_num, 6).value
    excel_armor_r = ws.cell(row_num, 7).value
    excel_weapon_1 = ws.cell(row_num, 8).value

    print(f"\nRow {row_num} EXCEL DATA:")
    print(f"  Name: {excel_name}")
    print(f"  Movement: {excel_move_off}/{excel_move_road}")
    print(f"  Armor: {excel_armor_f}/{excel_armor_s}/{excel_armor_r}")
    print(f"  Weapon_1: {excel_weapon_1}")

    # Look up in database
    cursor.execute("""
        SELECT id, name, movement_off_road, movement_road,
               armor_front, armor_side, armor_rear, weapon_1_id
        FROM bg_builder_vehicles
        WHERE name = ?
    """, (excel_name,))
    db_vehicle = cursor.fetchone()

    if db_vehicle:
        weapon_name = None
        if db_vehicle['weapon_1_id']:
            cursor.execute("SELECT weapon_name FROM bg_builder_weapons WHERE weapon_id = ?",
                          (db_vehicle['weapon_1_id'],))
            w = cursor.fetchone()
            if w:
                weapon_name = w['weapon_name']

        print(f"  DATABASE (bg_builder_vehicles ID {db_vehicle['id']}):")
        print(f"    Movement: {db_vehicle['movement_off_road']}/{db_vehicle['movement_road']}")
        print(f"    Armor: {db_vehicle['armor_front']}/{db_vehicle['armor_side']}/{db_vehicle['armor_rear']}")
        print(f"    Weapon_1: {weapon_name}")

        # VALIDATION
        match = True
        if excel_move_off != db_vehicle['movement_off_road']:
            print(f"  MISMATCH: Movement off-road {excel_move_off} != {db_vehicle['movement_off_road']}")
            match = False
        if excel_armor_f != db_vehicle['armor_front']:
            print(f"  MISMATCH: Armor front {excel_armor_f} != {db_vehicle['armor_front']}")
            match = False
        if excel_weapon_1 != weapon_name:
            print(f"  MISMATCH: Weapon_1 {excel_weapon_1} != {weapon_name}")
            match = False

        if match:
            print(f"  MATCH: Excel data matches bg_builder_vehicles")
    else:
        print(f"  NOT FOUND in bg_builder_vehicles")

conn.close()
