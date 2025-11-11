#!/usr/bin/env python3
"""Fix Marmon Herrington II A (20mm) record"""
import sqlite3
from pathlib import Path
from openpyxl import load_workbook

DB_PATH = Path(__file__).parent / "database" / "master_database.db"
EXPORT_PATH = Path(__file__).parent / "Vehicles_Tobruk_Torch_Export.xlsx"

# Find the row in spreadsheet
wb = load_workbook(EXPORT_PATH)
ws = wb.active

vehicle_data = None
for row_num in range(2, ws.max_row + 1):
    name = ws.cell(row_num, 1).value
    if name and '(20mm)' in name and 'Marmon' in name:
        vehicle_data = {
            'spreadsheet_name': name,
            'datacard_name': ws.cell(row_num, 2).value,
            'off_road_inches': ws.cell(row_num, 3).value,
            'road_inches': ws.cell(row_num, 4).value,
            'special_movement': ws.cell(row_num, 5).value,
            'armor_front': ws.cell(row_num, 6).value,
            'armor_side': ws.cell(row_num, 7).value,
            'armor_rear': ws.cell(row_num, 8).value,
            'weapon_1': ws.cell(row_num, 9).value,
            'weapon_2': ws.cell(row_num, 10).value,
            'weapon_3': ws.cell(row_num, 11).value,
            'weapon_4': ws.cell(row_num, 12).value,
            'mount_1': ws.cell(row_num, 13).value,
            'mount_2': ws.cell(row_num, 14).value,
            'mount_3': ws.cell(row_num, 15).value,
            'mount_4': ws.cell(row_num, 16).value,
            'ammo_1': ws.cell(row_num, 17).value,
            'ammo_2': ws.cell(row_num, 18).value,
            'ammo_3': ws.cell(row_num, 19).value,
            'ammo_4': ws.cell(row_num, 20).value,
            'armor_modifier': ws.cell(row_num, 21).value,
            'armor_side_schurzen': ws.cell(row_num, 22).value,
            'ss_hits': ws.cell(row_num, 23).value,
            'ss_transport_capacity': ws.cell(row_num, 24).value,
            'ss_special': ws.cell(row_num, 25).value,
            'year_range': ws.cell(row_num, 26).value,
            'vehicle_type': ws.cell(row_num, 27).value,
            'nation': ws.cell(row_num, 28).value,
            'dc_meta': ws.cell(row_num, 29).value,
            'source_battle': ws.cell(row_num, 30).value,
            'extraction_method': ws.cell(row_num, 31).value
        }
        print(f"Found in spreadsheet row {row_num}:")
        print(f"  Name: {name}")
        print(f"  Datacard name: {vehicle_data['datacard_name']}")
        break

if not vehicle_data:
    print("NOT FOUND in spreadsheet")
    exit(1)

# Update database
conn = sqlite3.connect(DB_PATH, timeout=60)
cursor = conn.cursor()

# This should be bg_builder_vehicles ID 345 (Marmon Herrington II A)
bg_builder_id = 345

# Get weapon_1_id
weapon_1_id = None
if vehicle_data['weapon_1']:
    cursor.execute("SELECT weapon_id FROM bg_builder_weapons WHERE weapon_name = ?",
                   (vehicle_data['weapon_1'],))
    result = cursor.fetchone()
    if result:
        weapon_1_id = result[0]

print(f"\nUpdating bg_builder_vehicles ID {bg_builder_id}...")

# Update bg_builder_vehicles
cursor.execute("""
    UPDATE bg_builder_vehicles
    SET movement_off_road = ?,
        movement_road = ?,
        movement_special = ?,
        armor_front = ?,
        armor_side = ?,
        armor_rear = ?,
        weapon_1_id = ?
    WHERE id = ?
""", (vehicle_data['off_road_inches'], vehicle_data['road_inches'],
      vehicle_data['special_movement'],
      vehicle_data['armor_front'], vehicle_data['armor_side'], vehicle_data['armor_rear'],
      weapon_1_id, bg_builder_id))

print(f"Updated bg_builder_vehicles")

# Check if exists in bg_reference_vehicles
cursor.execute("SELECT id FROM bg_reference_vehicles WHERE bg_builder_id = ?", (bg_builder_id,))
ref_result = cursor.fetchone()

if ref_result:
    # UPDATE
    print(f"Updating existing bg_reference_vehicles record...")
    cursor.execute("""
        UPDATE bg_reference_vehicles
        SET datacard_name = ?,
            weapon_2 = ?, weapon_3 = ?, weapon_4 = ?,
            mount_1 = ?, mount_2 = ?, mount_3 = ?, mount_4 = ?,
            ammo_1 = ?, ammo_2 = ?, ammo_3 = ?, ammo_4 = ?,
            armor_modifier = ?, armor_side_schurzen = ?,
            ss_hits = ?, ss_transport_capacity = ?, ss_special = ?,
            year_range = ?, vehicle_type = ?, nation = ?,
            dc_meta = ?, source_battle = ?, extraction_method = ?
        WHERE bg_builder_id = ?
    """, (vehicle_data['datacard_name'],
          vehicle_data['weapon_2'], vehicle_data['weapon_3'], vehicle_data['weapon_4'],
          vehicle_data['mount_1'], vehicle_data['mount_2'], vehicle_data['mount_3'], vehicle_data['mount_4'],
          vehicle_data['ammo_1'], vehicle_data['ammo_2'], vehicle_data['ammo_3'], vehicle_data['ammo_4'],
          vehicle_data['armor_modifier'], vehicle_data['armor_side_schurzen'],
          vehicle_data['ss_hits'], vehicle_data['ss_transport_capacity'], vehicle_data['ss_special'],
          vehicle_data['year_range'], vehicle_data['vehicle_type'], vehicle_data['nation'],
          vehicle_data['dc_meta'], vehicle_data['source_battle'], vehicle_data['extraction_method'],
          bg_builder_id))
    print(f"Updated bg_reference_vehicles")
else:
    # INSERT
    print(f"Inserting new bg_reference_vehicles record...")
    cursor.execute("""
        INSERT INTO bg_reference_vehicles
        (name, datacard_name, bg_builder_id,
         weapon_2, weapon_3, weapon_4,
         mount_1, mount_2, mount_3, mount_4,
         ammo_1, ammo_2, ammo_3, ammo_4,
         armor_modifier, armor_side_schurzen,
         ss_hits, ss_transport_capacity, ss_special,
         year_range, vehicle_type, nation,
         dc_meta, source_battle, extraction_method)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, ('Marmon Herrington II A', vehicle_data['datacard_name'], bg_builder_id,
          vehicle_data['weapon_2'], vehicle_data['weapon_3'], vehicle_data['weapon_4'],
          vehicle_data['mount_1'], vehicle_data['mount_2'], vehicle_data['mount_3'], vehicle_data['mount_4'],
          vehicle_data['ammo_1'], vehicle_data['ammo_2'], vehicle_data['ammo_3'], vehicle_data['ammo_4'],
          vehicle_data['armor_modifier'], vehicle_data['armor_side_schurzen'],
          vehicle_data['ss_hits'], vehicle_data['ss_transport_capacity'], vehicle_data['ss_special'],
          vehicle_data['year_range'], vehicle_data['vehicle_type'], vehicle_data['nation'],
          vehicle_data['dc_meta'], vehicle_data['source_battle'], vehicle_data['extraction_method']))
    print(f"Inserted into bg_reference_vehicles")

conn.commit()
conn.close()

print("\nFix complete!")
