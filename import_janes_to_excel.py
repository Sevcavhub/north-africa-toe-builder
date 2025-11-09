#!/usr/bin/env python3
"""
Import Jane's Guide Ammunition Data to Vehicles Manual Entry Form Excel

Strategy:
1. Load Jane's ammunition JSON data
2. Load Excel manual entry form
3. Match vehicles by name (fuzzy matching)
4. Update ammo_1-4 fields in Excel
5. Save updated Excel file
"""

import json
import openpyxl
from pathlib import Path
from difflib import SequenceMatcher

JANES_JSON = Path("janes_ammunition_v2.json")
EXCEL_FORM = Path("Resource Documents/Battlegroup Game/Vehicles Manual Entry Form.xlsx")
OUTPUT_EXCEL = Path("Resource Documents/Battlegroup Game/Vehicles Manual Entry Form - Updated.xlsx")


def normalize_name(name):
    """Normalize vehicle name for matching."""
    import re

    # Lowercase
    name = name.lower()

    # Remove punctuation
    name = re.sub(r'[^\w\s]', ' ', name)

    # Collapse whitespace
    name = re.sub(r'\s+', ' ', name).strip()

    # Common replacements
    replacements = {
        'mark ': 'mk ',
        ' tank': '',
        ' light': '',
        ' medium': '',
        ' heavy': '',
        ' infantry': '',
        ' cruiser': '',
    }

    for old, new in replacements.items():
        name = name.replace(old, new)

    return name


def similarity_ratio(str1, str2):
    """Calculate similarity ratio between two strings."""
    return SequenceMatcher(None, str1.lower(), str2.lower()).ratio()


def find_best_match(vehicle_name, excel_names, threshold=0.6):
    """
    Find best matching Excel entry for Jane's vehicle name.

    Returns: (excel_row_index, match_score) or None
    """
    normalized_target = normalize_name(vehicle_name)

    best_match = None
    best_score = threshold

    for row_idx, excel_name in excel_names.items():
        normalized_excel = normalize_name(excel_name)

        score = similarity_ratio(normalized_target, normalized_excel)

        if score > best_score:
            best_score = score
            best_match = (row_idx, score)

    return best_match


def main():
    """Main execution."""

    print("=" * 80)
    print("IMPORT JANE'S AMMUNITION DATA TO EXCEL VEHICLE FORM")
    print("=" * 80)

    # Load Jane's data
    print(f"\nLoading Jane's data: {JANES_JSON}")

    if not JANES_JSON.exists():
        print(f"ERROR: File not found: {JANES_JSON}")
        return

    with open(JANES_JSON, 'r') as f:
        janes_data = json.load(f)

    print(f"Loaded {len(janes_data)} vehicle entries from Jane's guide")

    # Filter out obvious non-vehicles (text fragments)
    filtered_janes = []
    for entry in janes_data:
        name = entry['vehicle_name']
        # Skip if name looks like a text fragment
        if any(x in name.lower() for x in ['carried', 'ammunition', 'rounds for', 'jane', 'world war']):
            continue
        # Skip very short names
        if len(name) < 5:
            continue
        filtered_janes.append(entry)

    print(f"Filtered to {len(filtered_janes)} likely vehicle entries")

    # Load Excel form
    print(f"\nLoading Excel form: {EXCEL_FORM}")

    if not EXCEL_FORM.exists():
        print(f"ERROR: File not found: {EXCEL_FORM}")
        return

    wb = openpyxl.load_workbook(EXCEL_FORM)
    sheet = wb.active

    # Get column indices
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}

    name_col = headers['name']
    ammo1_col = headers['ammo_1']
    ammo2_col = headers.get('ammo_2')
    ammo3_col = headers.get('ammo_3')
    ammo4_col = headers.get('ammo_4')

    print(f"Excel has {sheet.max_row - 1} vehicle rows")

    # Build Excel name index
    excel_names = {}
    for row in range(2, sheet.max_row + 1):
        name_cell = sheet.cell(row=row, column=name_col)
        if name_cell.value:
            excel_names[row] = name_cell.value

    print(f"Found {len(excel_names)} named vehicles in Excel")

    # Match and import
    print("\n" + "=" * 80)
    print("MATCHING JANE'S DATA TO EXCEL ENTRIES")
    print("=" * 80)

    matches = []
    no_matches = []

    for janes_entry in filtered_janes:
        vehicle_name = janes_entry['vehicle_name']
        ammo_count = janes_entry['ammunition_count']

        match = find_best_match(vehicle_name, excel_names, threshold=0.6)

        if match:
            row_idx, score = match
            excel_name = excel_names[row_idx]
            matches.append({
                'janes_name': vehicle_name,
                'excel_name': excel_name,
                'excel_row': row_idx,
                'ammo_count': ammo_count,
                'match_score': score
            })
        else:
            no_matches.append({
                'janes_name': vehicle_name,
                'ammo_count': ammo_count
            })

    print(f"\nMatching results:")
    print(f"  Matched: {len(matches)} vehicles")
    print(f"  No match: {len(no_matches)} vehicles")

    # Display matches
    if matches:
        print(f"\nMatched vehicles (sorted by confidence):\n")
        janes_header = "Jane's Name"
        print(f"{janes_header:45} | {'Excel Name':30} | {'Ammo':>5} | {'Score':>5}")
        print("-" * 95)

        for match in sorted(matches, key=lambda x: x['match_score'], reverse=True):
            print(f"{match['janes_name'][:45]:45} | {match['excel_name'][:30]:30} | "
                  f"{match['ammo_count']:5} | {match['match_score']*100:5.1f}%")

    # Ask for confirmation
    print(f"\n" + "=" * 80)
    print("IMPORT TO EXCEL")
    print("=" * 80)

    proceed = input(f"\nImport {len(matches)} ammunition values to Excel? (y/n): ")

    if proceed.lower() != 'y':
        print("Import cancelled.")
        return

    # Import data
    updates_made = 0

    for match in matches:
        row_idx = match['excel_row']
        ammo_count = match['ammo_count']

        # Get current ammo_1 value
        current_ammo = sheet.cell(row=row_idx, column=ammo1_col).value

        if current_ammo is None or current_ammo == '':
            # Update ammo_1 field
            sheet.cell(row=row_idx, column=ammo1_col, value=ammo_count)
            updates_made += 1
            print(f"  Updated: {match['excel_name']} -> ammo_1 = {ammo_count}")
        else:
            print(f"  Skipped: {match['excel_name']} (already has ammo: {current_ammo})")

    # Save Excel file
    print(f"\nSaving updated Excel file...")
    wb.save(OUTPUT_EXCEL)

    print(f"\n" + "=" * 80)
    print("IMPORT COMPLETE")
    print("=" * 80)
    print(f"\nUpdates made: {updates_made} vehicles")
    print(f"Saved to: {OUTPUT_EXCEL}")
    print(f"\nNo matches found for {len(no_matches)} vehicles:")

    for item in no_matches[:10]:
        print(f"  - {item['janes_name']} ({item['ammo_count']} rounds)")

    if len(no_matches) > 10:
        print(f"  ... and {len(no_matches) - 10} more")

    print(f"\nNext steps:")
    print(f"  1. Review: {OUTPUT_EXCEL}")
    print(f"  2. Manually add unmatched vehicles if needed")
    print(f"  3. Replace original file when satisfied")


if __name__ == "__main__":
    main()
