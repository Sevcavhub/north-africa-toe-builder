#!/usr/bin/env python3
"""
CLEAN SQL EXPORT - Tobruk/Torch Vehicles
NO JOINS - Pull fields ONLY from specified tables
"""
import sqlite3
from pathlib import Path
from openpyxl import load_workbook

DB_PATH = Path(__file__).parent / "database" / "master_database.db"
TEMPLATE_PATH = Path(__file__).parent / "Resource Documents" / "Battlegroup Game" / "Vehicles Manual Entry Form - Updated.xlsx"
OUTPUT_PATH = Path(__file__).parent / "Vehicles_Tobruk_Torch_Export.xlsx"
VEHICLE_IDS_PATH = Path(__file__).parent / "tobruk_torch_vehicle_ids.txt"

def export_vehicles():
    print("=" * 80)
    print("CLEAN SQL EXPORT - TOBRUK/TORCH VEHICLES")
    print("=" * 80)

    # Load Tobruk/Torch vehicle IDs
    with open(VEHICLE_IDS_PATH, 'r') as f:
        vehicle_ids = [int(line.strip()) for line in f if line.strip()]
    print(f"\nLoaded {len(vehicle_ids)} Tobruk/Torch vehicle IDs")

    # Load template
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # CLEAR ALL EXISTING DATA ROWS (keep row 1 header only)
    print("\nClearing existing data from template...")
    max_row = ws.max_row
    if max_row > 1:
        ws.delete_rows(2, max_row - 1)
    print(f"Cleared {max_row - 1} existing rows")

    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # Get ALL bg_builder_vehicles for these IDs
    vehicle_ids_str = ','.join(map(str, vehicle_ids))

    # Step 1: Get bg_builder data
    cursor.execute(f"""
        SELECT
            bgb.id,
            bgb.name,
            bgb.movement_off_road,
            bgb.movement_road,
            bgb.movement_special,
            bgb.armor_front,
            bgb.armor_side,
            bgb.armor_rear,
            bgb.weapon_1_id
        FROM bg_builder_vehicles bgb
        WHERE bgb.id IN ({vehicle_ids_str})
        ORDER BY bgb.name
    """)

    bg_vehicles = cursor.fetchall()
    print(f"Found {len(bg_vehicles)} bg_builder_vehicles")

    # Write to spreadsheet
    row_num = 2
    for bgv in bg_vehicles:
        # Get weapon_1 name if weapon_1_id exists
        weapon_1_name = None
        if bgv['weapon_1_id']:
            cursor.execute("SELECT weapon_name FROM bg_builder_weapons WHERE weapon_id = ?",
                          (bgv['weapon_1_id'],))
            w = cursor.fetchone()
            if w:
                weapon_1_name = w['weapon_name']

        # Get reference data if linkage exists
        cursor.execute("""
            SELECT weapon_2, weapon_3, weapon_4,
                   mount_1, mount_2, mount_3, mount_4,
                   ammo_1, ammo_2, ammo_3, ammo_4,
                   armor_modifier, armor_side_schurzen,
                   ss_hits, ss_transport_capacity, ss_special,
                   year_range, vehicle_type, nation,
                   dc_meta, source_battle, extraction_method
            FROM bg_reference_vehicles
            WHERE bg_builder_id = ?
            LIMIT 1
        """, (bgv['id'],))
        ref = cursor.fetchone()

        # Write BG Builder data (columns 1-8)
        ws.cell(row=row_num, column=1, value=bgv['name'])
        ws.cell(row=row_num, column=2, value=bgv['movement_off_road'])
        ws.cell(row=row_num, column=3, value=bgv['movement_road'])
        ws.cell(row=row_num, column=4, value=bgv['movement_special'])
        ws.cell(row=row_num, column=5, value=bgv['armor_front'])
        ws.cell(row=row_num, column=6, value=bgv['armor_side'])
        ws.cell(row=row_num, column=7, value=bgv['armor_rear'])
        ws.cell(row=row_num, column=8, value=weapon_1_name)

        # Write reference data (columns 9-30) if exists
        if ref:
            ws.cell(row=row_num, column=9, value=ref['weapon_2'])
            ws.cell(row=row_num, column=10, value=ref['weapon_3'])
            ws.cell(row=row_num, column=11, value=ref['weapon_4'])
            ws.cell(row=row_num, column=12, value=ref['mount_1'])
            ws.cell(row=row_num, column=13, value=ref['mount_2'])
            ws.cell(row=row_num, column=14, value=ref['mount_3'])
            ws.cell(row=row_num, column=15, value=ref['mount_4'])
            ws.cell(row=row_num, column=16, value=ref['ammo_1'])
            ws.cell(row=row_num, column=17, value=ref['ammo_2'])
            ws.cell(row=row_num, column=18, value=ref['ammo_3'])
            ws.cell(row=row_num, column=19, value=ref['ammo_4'])
            ws.cell(row=row_num, column=20, value=ref['armor_modifier'])
            ws.cell(row=row_num, column=21, value=ref['armor_side_schurzen'])
            ws.cell(row=row_num, column=22, value=ref['ss_hits'])
            ws.cell(row=row_num, column=23, value=ref['ss_transport_capacity'])
            ws.cell(row=row_num, column=24, value=ref['ss_special'])
            ws.cell(row=row_num, column=25, value=ref['year_range'])
            ws.cell(row=row_num, column=26, value=ref['vehicle_type'])
            ws.cell(row=row_num, column=27, value=ref['nation'])
            ws.cell(row=row_num, column=28, value=ref['dc_meta'])
            ws.cell(row=row_num, column=29, value=ref['source_battle'])
            ws.cell(row=row_num, column=30, value=ref['extraction_method'])

        row_num += 1

    wb.save(OUTPUT_PATH)
    conn.close()

    print(f"\nExport complete: {OUTPUT_PATH}")
    print(f"   {len(bg_vehicles)} vehicles exported")
    print(f"   Columns 1-8: BG Builder data ONLY")
    print(f"   Columns 9-30: bg_reference_vehicles data (where linked)")

if __name__ == '__main__':
    export_vehicles()
