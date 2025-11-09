import openpyxl
wb = openpyxl.load_workbook("Resource Documents/Battlegroup Game/Vehicles Manual Entry Form - Updated.xlsx")
ws = wb.active
print("Sheet:", ws.title)
print(f"Size: {ws.max_row} rows x {ws.max_column} cols\n")
print("Headers:")
for i in range(1, min(ws.max_column + 1, 30)):
    cell = ws.cell(1, i)
    if cell.value:
        print(f"  Col {i:2d}: {cell.value}")
