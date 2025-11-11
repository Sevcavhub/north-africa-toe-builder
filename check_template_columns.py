#!/usr/bin/env python3
"""Check Excel template column headers"""
from pathlib import Path
from openpyxl import load_workbook

TEMPLATE_PATH = Path(__file__).parent / "Resource Documents" / "Battlegroup Game" / "Vehicles Manual Entry Form - Updated.xlsx"

wb = load_workbook(TEMPLATE_PATH)
ws = wb.active

print("Template column headers:")
print("=" * 80)
for col_num in range(1, 50):
    header = ws.cell(row=1, column=col_num).value
    if header is None:
        break
    print(f"Column {col_num}: {header}")
