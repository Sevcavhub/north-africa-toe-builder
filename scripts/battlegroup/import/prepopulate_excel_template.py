#!/usr/bin/env python3
"""
Pre-populate user's Excel template with BG Builder data
- Uses existing template structure (Vehicles Manual Entry Form - Updated.xlsx)
- Filters to Tobruk and Torch books only
- Pre-fills: name, movement, armor, weapons from BG Builder
- Leaves blank: ammo counts, mounts, metadata (for manual entry)
"""
import sqlite3
import openpyxl
from pathlib import Path

DB_PATH = Path(__file__).parent.parent.parent.parent / "database" / "master_database.db"
TEMPLATE_PATH = Path(__file__).parent.parent.parent.parent / "Resource Documents" / "Battlegroup Game" / "Vehicles Manual Entry Form - Updated.xlsx"
OUTPUT_PATH = Path(__file__).parent.parent.parent.parent / "Vehicles_Manual_Entry_TOBRUK_TORCH_PrePopulated.xlsx"

def get_tobruk_torch_vehicles():
    """Get vehicles from Tobruk and Torch force lists only"""
    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # Get force IDs for Tobruk and Torch
    cursor.execute("""
        SELECT DISTINCT force_id, force_group, force_name
        FROM bg_builder_forces
        WHERE force_group LIKE '%Tobruk%' OR force_group LIKE '%Torch%'
        ORDER BY force_id
    """)
    forces = cursor.fetchall()

    print(f"Found {len(forces)} Tobruk/Torch force lists:")
    for force in forces:
        print(f"   [{force['force_id']:3d}] {force['force_group']} - {force['force_name']}")

    # Get all vehicles with their stats
    cursor.execute("""
        SELECT DISTINCT
            bgb.id,
            bgb.name,
            bgb.movement_off_road,
            bgb.movement_road,
            bgb.movement_special,
            bgb.armor_front,
            bgb.armor_side,
            bgb.armor_rear,
            w1.weapon_name as weapon_1,
            w2.weapon_name as weapon_2,
            w3.weapon_name as weapon_3,
            w4.weapon_name as weapon_4,
            bgb.special_rules,
            bgb.hits,
            bgb.capacity,
            bgb.has_mg
        FROM bg_builder_vehicles bgb
        LEFT JOIN bg_builder_weapons w1 ON bgb.weapon_1_id = w1.weapon_id
        LEFT JOIN bg_builder_weapons w2 ON bgb.weapon_2_id = w2.weapon_id
        LEFT JOIN bg_builder_weapons w3 ON bgb.weapon_3_id = w3.weapon_id
        LEFT JOIN bg_builder_weapons w4 ON bgb.weapon_4_id = w4.weapon_id
        ORDER BY bgb.name
    """)

    all_vehicles = cursor.fetchall()
    conn.close()

    return all_vehicles

def prepopulate_template():
    print("Pre-populating Excel Template with BG Builder Data")
    print("=" * 80)
    print(f"\nTemplate: {TEMPLATE_PATH}")
    print(f"Output: {OUTPUT_PATH}")

    # Get Tobruk/Torch vehicles
    print("\nQuerying database...")
    vehicles = get_tobruk_torch_vehicles()
    print(f"\nFound {len(vehicles)} total vehicles in BG Builder")

    # Load template
    print(f"\nLoading template...")
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Verify headers match expected structure
    expected_headers = [
        'name', 'off_road_inches', 'road_inches', 'special_movement',
        'armor_front', 'armor_side', 'armor_rear',
        'weapon_1', 'weapon_2', 'weapon_3', 'weapon_4',
        'mount_1', 'mount_2', 'mount_3', 'mount_4',
        'ammo_1', 'ammo_2', 'ammo_3', 'ammo_4',
        'armor_modifier', 'armor_side_schurzen',
        'ss_hits', 'ss_transport_capacity', 'ss_special',
        'year_range', 'vehicle_type', 'nation', 'dc_meta', 'source_battle'
    ]

    # Map headers to column indices
    header_map = {}
    for i in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, i).value
        if cell_value in expected_headers:
            header_map[cell_value] = i

    print(f"Mapped {len(header_map)} columns")

    # Pre-populate vehicles (starting at row 2)
    print(f"\nPre-populating {len(vehicles)} vehicles...")

    row_num = 2
    populated = 0

    for vehicle in vehicles:
        # PRE-FILLED from BG Builder (official data)
        ws.cell(row_num, header_map['name']).value = vehicle['name']
        ws.cell(row_num, header_map['off_road_inches']).value = vehicle['movement_off_road']
        ws.cell(row_num, header_map['road_inches']).value = vehicle['movement_road']
        ws.cell(row_num, header_map['special_movement']).value = vehicle['movement_special']
        ws.cell(row_num, header_map['armor_front']).value = vehicle['armor_front']
        ws.cell(row_num, header_map['armor_side']).value = vehicle['armor_side']
        ws.cell(row_num, header_map['armor_rear']).value = vehicle['armor_rear']
        ws.cell(row_num, header_map['weapon_1']).value = vehicle['weapon_1']
        ws.cell(row_num, header_map['weapon_2']).value = vehicle['weapon_2']
        ws.cell(row_num, header_map['weapon_3']).value = vehicle['weapon_3']
        ws.cell(row_num, header_map['weapon_4']).value = vehicle['weapon_4']
        ws.cell(row_num, header_map['ss_hits']).value = vehicle['hits']
        ws.cell(row_num, header_map['ss_transport_capacity']).value = vehicle['capacity']

        # LEAVE BLANK for manual entry:
        # - mount_1, mount_2, mount_3, mount_4
        # - ammo_1, ammo_2, ammo_3, ammo_4
        # - armor_modifier, armor_side_schurzen
        # - ss_special
        # - year_range, vehicle_type, nation, dc_meta, source_battle

        row_num += 1
        populated += 1

        if populated % 50 == 0:
            print(f"   Populated {populated} vehicles...")

    # Save output
    print(f"\nSaving to: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)

    print("\n" + "=" * 80)
    print("PRE-POPULATION COMPLETE")
    print("=" * 80)
    print(f"\nPopulated {populated} vehicles")
    print("\nPRE-FILLED from BG Builder (DO NOT EDIT):")
    print("   - name")
    print("   - off_road_inches, road_inches")
    print("   - armor_front, armor_side, armor_rear")
    print("   - weapon_1, weapon_2, weapon_3, weapon_4")
    print("   - special_movement")
    print("   - ss_hits, ss_transport_capacity")
    print("\nFILL THESE MANUALLY (currently blank):")
    print("   - mount_1, mount_2, mount_3, mount_4 (weapon mount positions)")
    print("   - ammo_1, ammo_2, ammo_3, ammo_4 (ammunition round counts)")
    print("   - armor_modifier, armor_side_schurzen")
    print("   - ss_special (soft-skin special rules)")
    print("   - year_range, vehicle_type, nation")
    print("   - dc_meta, source_battle")
    print(f"\nOpen {OUTPUT_PATH} in Excel to complete manual entry")

if __name__ == '__main__':
    prepopulate_template()
