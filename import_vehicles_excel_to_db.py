#!/usr/bin/env python3
"""
Import Vehicles Manual Entry Form to bg_reference_vehicles Database

Reads Excel manual entry form and imports all vehicle data into the database.
Updates existing records or inserts new ones based on name+nation uniqueness.
"""

import openpyxl
import sqlite3
from pathlib import Path
from datetime import datetime

EXCEL_FILE = Path("Resource Documents/Battlegroup Game/Vehicles Manual Entry Form.xlsx")
DATABASE_PATH = Path("database/master_database.db")


def read_excel_vehicles(excel_path):
    """Read vehicle data from Excel manual entry form."""

    print(f"\nReading Excel file: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Get headers from first row
    headers = {}
    for cell in sheet[1]:
        if cell.value:
            headers[cell.value] = cell.column

    print(f"Found {len(headers)} columns: {list(headers.keys())}")

    # Read all vehicle rows
    vehicles = []
    for row_idx in range(2, sheet.max_row + 1):
        # Check if row has data (name field must exist)
        name_col = headers.get('name')
        if not name_col:
            print("ERROR: 'name' column not found in Excel")
            return []

        name_cell = sheet.cell(row=row_idx, column=name_col)
        if not name_cell.value:
            continue  # Skip empty rows

        # Build vehicle record from all columns
        vehicle = {}
        for header, col_idx in headers.items():
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            vehicle[header] = cell_value

        vehicles.append(vehicle)

    print(f"Read {len(vehicles)} vehicle records from Excel\n")

    return vehicles


def import_to_database(vehicles, conn):
    """Import vehicles to bg_reference_vehicles table."""

    if not vehicles:
        print("No vehicles to import")
        return

    cursor = conn.cursor()

    # Get current table schema
    cursor.execute("PRAGMA table_info(bg_reference_vehicles)")
    db_columns = {row[1] for row in cursor.fetchall()}

    print(f"Database has {len(db_columns)} columns")

    # Get Excel column names from first vehicle
    excel_columns = set(vehicles[0].keys())

    print(f"Excel has {len(excel_columns)} columns")

    # Find matching columns
    matching_columns = excel_columns & db_columns
    print(f"Matching columns: {len(matching_columns)}")
    print(f"Columns: {sorted(matching_columns)}\n")

    # Import each vehicle
    updates = 0
    inserts = 0
    errors = 0

    for vehicle in vehicles:
        try:
            name = vehicle.get('name')
            nation = vehicle.get('nation')

            if not name:
                print(f"  SKIP: Vehicle missing name")
                errors += 1
                continue

            # Check if vehicle already exists (by name + nation)
            cursor.execute(
                "SELECT id FROM bg_reference_vehicles WHERE name = ? AND nation = ?",
                (name, nation)
            )
            existing = cursor.fetchone()

            # Build column list and values for matching columns only
            columns = []
            values = []

            for col in matching_columns:
                if col in vehicle and vehicle[col] is not None:
                    columns.append(col)
                    values.append(vehicle[col])

            if existing:
                # UPDATE existing record
                vehicle_id = existing[0]

                # Build UPDATE statement
                set_clause = ", ".join([f"{col} = ?" for col in columns])
                values.append(vehicle_id)  # WHERE id = ?

                sql = f"UPDATE bg_reference_vehicles SET {set_clause} WHERE id = ?"
                cursor.execute(sql, values)

                updates += 1
                print(f"  UPDATE: {name} ({nation}) - ID {vehicle_id}")

            else:
                # INSERT new record
                placeholders = ", ".join(["?" for _ in columns])
                columns_str = ", ".join(columns)

                sql = f"INSERT INTO bg_reference_vehicles ({columns_str}) VALUES ({placeholders})"
                cursor.execute(sql, values)

                inserts += 1
                print(f"  INSERT: {name} ({nation}) - NEW")

        except Exception as e:
            errors += 1
            print(f"  ERROR: {vehicle.get('name', 'UNKNOWN')} - {e}")

    conn.commit()

    print(f"\n" + "=" * 80)
    print("IMPORT COMPLETE")
    print("=" * 80)
    print(f"Updated: {updates} records")
    print(f"Inserted: {inserts} records")
    print(f"Errors: {errors} records")
    print(f"Total processed: {len(vehicles)} records\n")


def main():
    """Main execution."""

    print("=" * 80)
    print("IMPORT VEHICLES MANUAL ENTRY FORM TO DATABASE")
    print("=" * 80)

    # Check files exist
    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        return

    if not DATABASE_PATH.exists():
        print(f"ERROR: Database not found: {DATABASE_PATH}")
        return

    # Read Excel data
    vehicles = read_excel_vehicles(EXCEL_FILE)

    if not vehicles:
        print("No vehicle data found in Excel")
        return

    # Display sample
    print("Sample vehicles to import:\n")
    print(f"{'Name':30} | {'Nation':10} | {'Off-Road':>8} | {'Road':>8} | {'Weapon 1':20}")
    print("-" * 90)

    for vehicle in vehicles[:10]:
        name = vehicle.get('name', '')[:30]
        nation = vehicle.get('nation', '')[:10]
        off_road = vehicle.get('off_road_inches', '')
        road = vehicle.get('road_inches', '')
        weapon = vehicle.get('weapon_1', '')[:20]

        print(f"{name:30} | {nation:10} | {str(off_road):>8} | {str(road):>8} | {weapon:20}")

    if len(vehicles) > 10:
        print(f"... and {len(vehicles) - 10} more")

    # Auto-confirm import
    print(f"\n" + "=" * 80)
    print(f"Importing {len(vehicles)} vehicles to bg_reference_vehicles...")
    print("=" * 80)

    # Connect to database and import
    conn = sqlite3.connect(DATABASE_PATH)

    try:
        import_to_database(vehicles, conn)
    finally:
        conn.close()

    print("Database connection closed")


if __name__ == "__main__":
    main()
