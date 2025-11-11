#!/usr/bin/env python3
"""
CAREFUL IMPORT - User's manually updated vehicle data
With full backups and validation
"""
import sqlite3
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

DB_PATH = Path(__file__).parent / "database" / "master_database.db"
EXPORT_PATH = Path(__file__).parent / "Vehicles_Tobruk_Torch_Export.xlsx"
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

def backup_tables(cursor):
    """Create timestamped backups of both tables"""
    print("=" * 80)
    print("STEP 1: CREATING BACKUPS")
    print("=" * 80)

    # Backup bg_reference_vehicles
    backup_ref = f"bg_reference_vehicles_backup_{timestamp}"
    cursor.execute(f"CREATE TABLE {backup_ref} AS SELECT * FROM bg_reference_vehicles")
    cursor.execute(f"SELECT COUNT(*) FROM {backup_ref}")
    count_ref = cursor.fetchone()[0]
    print(f"\nOK Backed up bg_reference_vehicles -> {backup_ref} ({count_ref} rows)")

    # Backup bg_builder_vehicles
    backup_builder = f"bg_builder_vehicles_backup_{timestamp}"
    cursor.execute(f"CREATE TABLE {backup_builder} AS SELECT * FROM bg_builder_vehicles")
    cursor.execute(f"SELECT COUNT(*) FROM {backup_builder}")
    count_builder = cursor.fetchone()[0]
    print(f"OK Backed up bg_builder_vehicles -> {backup_builder} ({count_builder} rows)")

    return backup_ref, backup_builder

def add_datacard_name_column(cursor):
    """Add datacard_name column to bg_reference_vehicles if not exists"""
    print("\n" + "=" * 80)
    print("STEP 2: ADDING datacard_name COLUMN")
    print("=" * 80)

    cursor.execute("PRAGMA table_info(bg_reference_vehicles)")
    columns = [col[1] for col in cursor.fetchall()]

    if 'datacard_name' not in columns:
        cursor.execute("ALTER TABLE bg_reference_vehicles ADD COLUMN datacard_name TEXT")
        print("\nOK Added datacard_name column to bg_reference_vehicles")
    else:
        print("\nOK datacard_name column already exists")

def load_spreadsheet_data(filepath):
    """Load all vehicle data from spreadsheet"""
    print("\n" + "=" * 80)
    print("STEP 3: LOADING SPREADSHEET DATA")
    print("=" * 80)

    wb = load_workbook(filepath)
    ws = wb.active

    vehicles = []
    for row_num in range(2, ws.max_row + 1):
        name = ws.cell(row_num, 1).value
        if not name:
            continue

        vehicle = {
            'name': name,
            'datacard_name': ws.cell(row_num, 2).value,
            'off_road_inches': ws.cell(row_num, 3).value,
            'road_inches': ws.cell(row_num, 4).value,
            'special_movement': ws.cell(row_num, 5).value,
            'armor_front': ws.cell(row_num, 6).value,
            'armor_side': ws.cell(row_num, 7).value,
            'armor_rear': ws.cell(row_num, 8).value,
            'weapon_1': ws.cell(row_num, 9).value,
            'weapon_2': ws.cell(row_num, 10).value,
            'weapon_3': ws.cell(row_num, 11).value,
            'weapon_4': ws.cell(row_num, 12).value,
            'mount_1': ws.cell(row_num, 13).value,
            'mount_2': ws.cell(row_num, 14).value,
            'mount_3': ws.cell(row_num, 15).value,
            'mount_4': ws.cell(row_num, 16).value,
            'ammo_1': ws.cell(row_num, 17).value,
            'ammo_2': ws.cell(row_num, 18).value,
            'ammo_3': ws.cell(row_num, 19).value,
            'ammo_4': ws.cell(row_num, 20).value,
            'armor_modifier': ws.cell(row_num, 21).value,
            'armor_side_schurzen': ws.cell(row_num, 22).value,
            'ss_hits': ws.cell(row_num, 23).value,
            'ss_transport_capacity': ws.cell(row_num, 24).value,
            'ss_special': ws.cell(row_num, 25).value,
            'year_range': ws.cell(row_num, 26).value,
            'vehicle_type': ws.cell(row_num, 27).value,
            'nation': ws.cell(row_num, 28).value,
            'dc_meta': ws.cell(row_num, 29).value,
            'source_battle': ws.cell(row_num, 30).value,
            'extraction_method': ws.cell(row_num, 31).value
        }
        vehicles.append(vehicle)

    print(f"\nOK Loaded {len(vehicles)} vehicles from spreadsheet")
    return vehicles

def update_bg_builder_vehicles(cursor, vehicles):
    """Update bg_builder_vehicles with corrected data from columns 3-9"""
    print("\n" + "=" * 80)
    print("STEP 4: UPDATING bg_builder_vehicles")
    print("=" * 80)

    updated_count = 0
    errors = []

    for v in vehicles:
        # Find bg_builder_vehicles entry by name
        cursor.execute("SELECT id FROM bg_builder_vehicles WHERE name = ?", (v['name'],))
        result = cursor.fetchone()

        if not result:
            errors.append(f"NOT FOUND in bg_builder_vehicles: {v['name']}")
            continue

        bg_id = result[0]

        # Get weapon_1_id from weapon name
        weapon_1_id = None
        if v['weapon_1']:
            cursor.execute("SELECT weapon_id FROM bg_builder_weapons WHERE weapon_name = ?", (v['weapon_1'],))
            w_result = cursor.fetchone()
            if w_result:
                weapon_1_id = w_result[0]

        # Update bg_builder_vehicles
        cursor.execute("""
            UPDATE bg_builder_vehicles
            SET movement_off_road = ?,
                movement_road = ?,
                movement_special = ?,
                armor_front = ?,
                armor_side = ?,
                armor_rear = ?,
                weapon_1_id = ?
            WHERE id = ?
        """, (v['off_road_inches'], v['road_inches'], v['special_movement'],
              v['armor_front'], v['armor_side'], v['armor_rear'],
              weapon_1_id, bg_id))

        updated_count += 1

    print(f"\nOK Updated {updated_count} vehicles in bg_builder_vehicles")
    if errors:
        print(f"\nErrors ({len(errors)}):")
        for err in errors[:10]:
            print(f"  {err}")

def update_bg_reference_vehicles(cursor, vehicles):
    """Update bg_reference_vehicles with all data including datacard_name"""
    print("\n" + "=" * 80)
    print("STEP 5: UPDATING bg_reference_vehicles")
    print("=" * 80)

    updated_count = 0
    inserted_count = 0
    errors = []

    for v in vehicles:
        # Find bg_builder_id
        cursor.execute("SELECT id FROM bg_builder_vehicles WHERE name = ?", (v['name'],))
        result = cursor.fetchone()

        if not result:
            errors.append(f"NOT FOUND in bg_builder_vehicles: {v['name']}")
            continue

        bg_builder_id = result[0]

        # Check if exists in bg_reference_vehicles
        cursor.execute("SELECT id FROM bg_reference_vehicles WHERE bg_builder_id = ?", (bg_builder_id,))
        ref_result = cursor.fetchone()

        if ref_result:
            # UPDATE existing
            cursor.execute("""
                UPDATE bg_reference_vehicles
                SET datacard_name = ?,
                    weapon_1 = ?, weapon_2 = ?, weapon_3 = ?, weapon_4 = ?,
                    mount_1 = ?, mount_2 = ?, mount_3 = ?, mount_4 = ?,
                    ammo_1 = ?, ammo_2 = ?, ammo_3 = ?, ammo_4 = ?,
                    armor_modifier = ?, armor_side_schurzen = ?,
                    ss_hits = ?, ss_transport_capacity = ?, ss_special = ?,
                    year_range = ?, vehicle_type = ?, nation = ?,
                    dc_meta = ?, source_battle = ?, extraction_method = ?
                WHERE bg_builder_id = ?
            """, (v['datacard_name'],
                  v['weapon_1'], v['weapon_2'], v['weapon_3'], v['weapon_4'],
                  v['mount_1'], v['mount_2'], v['mount_3'], v['mount_4'],
                  v['ammo_1'], v['ammo_2'], v['ammo_3'], v['ammo_4'],
                  v['armor_modifier'], v['armor_side_schurzen'],
                  v['ss_hits'], v['ss_transport_capacity'], v['ss_special'],
                  v['year_range'], v['vehicle_type'], v['nation'],
                  v['dc_meta'], v['source_battle'], v['extraction_method'],
                  bg_builder_id))
            updated_count += 1
        else:
            # INSERT new (vehicle not previously in bg_reference_vehicles)
            cursor.execute("""
                INSERT INTO bg_reference_vehicles
                (name, datacard_name, bg_builder_id,
                 weapon_1, weapon_2, weapon_3, weapon_4,
                 mount_1, mount_2, mount_3, mount_4,
                 ammo_1, ammo_2, ammo_3, ammo_4,
                 armor_modifier, armor_side_schurzen,
                 ss_hits, ss_transport_capacity, ss_special,
                 year_range, vehicle_type, nation,
                 dc_meta, source_battle, extraction_method)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (v['name'], v['datacard_name'], bg_builder_id,
                  v['weapon_1'], v['weapon_2'], v['weapon_3'], v['weapon_4'],
                  v['mount_1'], v['mount_2'], v['mount_3'], v['mount_4'],
                  v['ammo_1'], v['ammo_2'], v['ammo_3'], v['ammo_4'],
                  v['armor_modifier'], v['armor_side_schurzen'],
                  v['ss_hits'], v['ss_transport_capacity'], v['ss_special'],
                  v['year_range'], v['vehicle_type'], v['nation'],
                  v['dc_meta'], v['source_battle'], v['extraction_method']))
            inserted_count += 1

    print(f"\nOK Updated {updated_count} existing vehicles")
    print(f"OK Inserted {inserted_count} new vehicles")
    if errors:
        print(f"\nErrors ({len(errors)}):")
        for err in errors[:10]:
            print(f"  {err}")

def validate_import(cursor):
    """Validate the import was successful"""
    print("\n" + "=" * 80)
    print("STEP 6: VALIDATION")
    print("=" * 80)

    # Check datacard_name populated
    cursor.execute("SELECT COUNT(*) FROM bg_reference_vehicles WHERE datacard_name IS NOT NULL")
    count = cursor.fetchone()[0]
    print(f"\nOK {count} vehicles have datacard_name")

    # Sample check
    cursor.execute("""
        SELECT bgb.name, bgb.armor_front, bgb.weapon_1_id, ref.datacard_name
        FROM bg_builder_vehicles bgb
        LEFT JOIN bg_reference_vehicles ref ON ref.bg_builder_id = bgb.id
        WHERE bgb.name = '1 tonne SdKfz 10'
    """)
    result = cursor.fetchone()
    if result:
        print(f"\nSample validation (1 tonne SdKfz 10):")
        print(f"  bg_builder name: {result[0]}")
        print(f"  bg_builder armor: {result[1]}")
        print(f"  datacard_name: {result[3]}")

def main():
    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        # Step 1: Backup
        backup_ref, backup_builder = backup_tables(cursor)
        conn.commit()

        # Step 2: Add column
        add_datacard_name_column(cursor)
        conn.commit()

        # Step 3: Load spreadsheet
        vehicles = load_spreadsheet_data(EXPORT_PATH)

        # Step 4: Update bg_builder_vehicles
        update_bg_builder_vehicles(cursor, vehicles)
        conn.commit()

        # Step 5: Update bg_reference_vehicles
        update_bg_reference_vehicles(cursor, vehicles)
        conn.commit()

        # Step 6: Validate
        validate_import(cursor)

        print("\n" + "=" * 80)
        print("IMPORT COMPLETE")
        print("=" * 80)
        print(f"\nBackup tables created:")
        print(f"  {backup_ref}")
        print(f"  {backup_builder}")
        print(f"\nTo restore if needed:")
        print(f"  DROP TABLE bg_reference_vehicles;")
        print(f"  ALTER TABLE {backup_ref} RENAME TO bg_reference_vehicles;")

    except Exception as e:
        conn.rollback()
        print(f"\nERROR ERROR: {e}")
        print("Transaction rolled back - no changes made")
        raise
    finally:
        conn.close()

if __name__ == '__main__':
    main()
