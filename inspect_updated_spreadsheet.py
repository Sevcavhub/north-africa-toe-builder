#!/usr/bin/env python3
"""Inspect the user's updated spreadsheet to understand changes"""
from pathlib import Path
from openpyxl import load_workbook

EXPORT_PATH = Path(__file__).parent / "Vehicles_Tobruk_Torch_Export.xlsx"

wb = load_workbook(EXPORT_PATH)
ws = wb.active

print("=" * 80)
print("SPREADSHEET STRUCTURE ANALYSIS")
print("=" * 80)

# Check headers
print("\nHEADER ROW (Row 1):")
headers = []
for col in range(1, 40):
    val = ws.cell(1, col).value
    if val:
        headers.append((col, val))
        print(f"  Col {col:2d}: {val}")
    else:
        break

print(f"\nTotal columns: {len(headers)}")

# Check for new columns beyond original 30
if len(headers) > 30:
    print(f"\n*** NEW COLUMNS DETECTED (beyond original 30): ***")
    for col_num, header in headers[30:]:
        print(f"  Col {col_num}: {header}")

# Sample first 3 data rows to see structure
print("\n" + "=" * 80)
print("SAMPLE DATA (First 3 rows):")
print("=" * 80)

for row in range(2, 5):
    name = ws.cell(row, 1).value
    if name:
        print(f"\nRow {row}: {name}")
        # Show first 10 columns
        for col in range(1, min(11, len(headers)+1)):
            val = ws.cell(row, col).value
            if col <= len(headers):
                print(f"  {headers[col-1][1]:20s}: {val}")

        # Show any new columns
        if len(headers) > 30:
            print("  NEW FIELDS:")
            for col_num, header in headers[30:]:
                val = ws.cell(row, col_num).value
                print(f"    {header}: {val}")
